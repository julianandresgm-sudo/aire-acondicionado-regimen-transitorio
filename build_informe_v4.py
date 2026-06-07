from __future__ import annotations

import csv
import math
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from statistics import mean, stdev
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE_DOCX = ROOT / "ensayo_rubatex_apa_mejorado_v3.docx"
OUTPUT_DOCX = ROOT / "informe_tecnico_rubatex_APA_final_v4.docx"
DATA_XLSX = ROOT / "1 prueba con espesor de 2mm.xlsx"
CHART_PNG = ROOT / "figura3_regimen_cuasiestacionario.png"


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
M_NS = "http://schemas.openxmlformats.org/officeDocument/2006/math"
ET.register_namespace("w", W_NS)
ET.register_namespace("m", M_NS)


def qn(tag: str) -> str:
    return f"{{{W_NS}}}{tag}"


def mn(tag: str) -> str:
    return f"{{{M_NS}}}{tag}"


def load_temperature_rows() -> list[dict[str, float | datetime]]:
    wb = load_workbook(DATA_XLSX, data_only=True, read_only=True)
    ws = wb["Temperaturas"]
    rows: list[dict[str, float | datetime]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        ts = row[0]
        if not isinstance(ts, datetime):
            continue
        if not all(isinstance(row[idx], (int, float)) for idx in (1, 2, 6)):
            continue
        rows.append(
            {
                "timestamp": ts,
                "amb": float(row[1]),
                "tc": float(row[2]),
                "ts": float(row[6]),
            }
        )
    return rows


def linfit_minutes(values: list[float], minutes: list[float]) -> float:
    xm = sum(minutes) / len(minutes)
    ym = sum(values) / len(values)
    denom = sum((x - xm) ** 2 for x in minutes)
    return sum((x - xm) * (y - ym) for x, y in zip(minutes, values)) / denom


def stability_stats(rows: list[dict[str, float | datetime]], n: int = 63) -> dict[str, dict[str, float | str]]:
    stable = rows[-n:]
    t0 = stable[0]["timestamp"]
    minutes = [((r["timestamp"] - t0).total_seconds() / 60.0) for r in stable]  # type: ignore[operator]
    out: dict[str, dict[str, float | str]] = {}
    for key, label in [("tc", "T_c"), ("ts", "T_s"), ("amb", "T_∞")]:
        vals = [float(r[key]) for r in stable]
        out[key] = {
            "label": label,
            "mean": mean(vals),
            "sd": stdev(vals),
            "min": min(vals),
            "max": max(vals),
            "slope": linfit_minutes(vals, minutes),
        }
    out["window"] = {
        "start": stable[0]["timestamp"].strftime("%H:%M:%S"),  # type: ignore[union-attr]
        "end": stable[-1]["timestamp"].strftime("%H:%M:%S"),  # type: ignore[union-attr]
        "n": n,
        "duration": (stable[-1]["timestamp"] - stable[0]["timestamp"]).total_seconds() / 60.0,  # type: ignore[operator]
    }
    return out


def write_chart_with_powershell(rows: list[dict[str, float | datetime]], stats: dict[str, dict[str, float | str]]) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        csv_path = tmpdir / "chart_data.csv"
        ps_path = tmpdir / "draw_chart.ps1"
        t0 = rows[0]["timestamp"]
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["minute", "amb", "tc", "ts", "stable"])
            stable_start = len(rows) - 63
            for i, row in enumerate(rows):
                minute = (row["timestamp"] - t0).total_seconds() / 60.0  # type: ignore[operator]
                writer.writerow([minute, row["amb"], row["tc"], row["ts"], 1 if i >= stable_start else 0])

        ps_path.write_text(
            r'''
param([string]$CsvPath, [string]$OutPath)
Add-Type -AssemblyName System.Drawing
$data = Import-Csv -Path $CsvPath
$W = 2200; $H = 1350
$bmp = New-Object System.Drawing.Bitmap($W, $H)
$g = [System.Drawing.Graphics]::FromImage($bmp)
$g.SmoothingMode = [System.Drawing.Drawing2D.SmoothingMode]::AntiAlias
$g.TextRenderingHint = [System.Drawing.Text.TextRenderingHint]::ClearTypeGridFit
$white = [System.Drawing.Color]::FromArgb(255,255,255)
$g.Clear($white)
$fontTitle = New-Object System.Drawing.Font("Arial", 34, [System.Drawing.FontStyle]::Bold)
$font = New-Object System.Drawing.Font("Arial", 22, [System.Drawing.FontStyle]::Regular)
$fontSmall = New-Object System.Drawing.Font("Arial", 18, [System.Drawing.FontStyle]::Regular)
$fontBold = New-Object System.Drawing.Font("Arial", 20, [System.Drawing.FontStyle]::Bold)
$black = [System.Drawing.Brushes]::Black
$axisPen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(40,40,40), 3)
$gridPen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(220,220,220), 1)
$dashPen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(110,110,110), 2)
$dashPen.DashStyle = [System.Drawing.Drawing2D.DashStyle]::Dash
$bluePen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(35,105,180), 5)
$orangePen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(230,120,25), 5)
$greenPen = New-Object System.Drawing.Pen([System.Drawing.Color]::FromArgb(45,155,70), 5)
$shadeBrush = New-Object System.Drawing.SolidBrush([System.Drawing.Color]::FromArgb(55,70,135,190))
$stableBrush = New-Object System.Drawing.SolidBrush([System.Drawing.Color]::FromArgb(18,45,155,70))

function X($m, $left, $plotW, $maxM) { return [float]($left + ($m / $maxM) * $plotW) }
function Y($v, $top, $plotH, $minV, $maxV) { return [float]($top + (($maxV - $v) / ($maxV - $minV)) * $plotH) }
function Draw-Series($key, $pen, $left, $top, $plotW, $plotH, $maxM, $minV, $maxV) {
  $pts = New-Object System.Collections.Generic.List[System.Drawing.PointF]
  foreach($r in $data){
    $pts.Add([System.Drawing.PointF]::new((X ([double]$r.minute) $left $plotW $maxM), (Y ([double]$r.$key) $top $plotH $minV $maxV)))
  }
  $g.DrawLines($pen, $pts.ToArray())
}

$left=135; $top=135; $plotW=1880; $plotH=520
$maxM = (($data | Select-Object -Last 1).minute -as [double])
$minV=28; $maxV=56
$g.DrawString("Figura 3. Evolucion termica y validacion de la ventana cuasi-estacionaria", $fontTitle, $black, 120, 45)
for($yv=28; $yv -le 56; $yv+=4){
  $yy=Y $yv $top $plotH $minV $maxV
  $g.DrawLine($gridPen, $left, $yy, $left+$plotW, $yy)
  $g.DrawString("$yv", $fontSmall, $black, 72, $yy-13)
}
for($xm=0; $xm -le 8; $xm+=1){
  $xx=X $xm $left $plotW $maxM
  $g.DrawLine($gridPen, $xx, $top, $xx, $top+$plotH)
  $g.DrawString("$xm", $fontSmall, $black, $xx-8, $top+$plotH+18)
}
$stableFirst = ($data | Where-Object stable -eq 1 | Select-Object -First 1)
$stableLast = ($data | Select-Object -Last 1)
$sx0 = X ([double]$stableFirst.minute) $left $plotW $maxM
$sx1 = X ([double]$stableLast.minute) $left $plotW $maxM
$g.FillRectangle($shadeBrush, $sx0, $top, $sx1-$sx0, $plotH)
$g.DrawLine($dashPen, $sx0, $top, $sx0, $top+$plotH)
$g.DrawString("ventana N=63", $fontBold, [System.Drawing.Brushes]::DarkSlateGray, $sx0+18, $top+18)
Draw-Series "amb" $bluePen $left $top $plotW $plotH $maxM $minV $maxV
Draw-Series "tc" $orangePen $left $top $plotW $plotH $maxM $minV $maxV
Draw-Series "ts" $greenPen $left $top $plotW $plotH $maxM $minV $maxV
$g.DrawRectangle($axisPen, $left, $top, $plotW, $plotH)
$g.DrawString("Tiempo desde el inicio de la prueba (min)", $font, $black, 820, $top+$plotH+52)
$g.TranslateTransform(35, 470); $g.RotateTransform(-90); $g.DrawString("Temperatura (C)", $font, $black, 0, 0); $g.ResetTransform()
$legendY=760; $legendX=145
$g.DrawLine($bluePen, $legendX, $legendY, $legendX+80, $legendY); $g.DrawString("Tinf ambiente", $font, $black, $legendX+95, $legendY-18)
$g.DrawLine($orangePen, $legendX+390, $legendY, $legendX+470, $legendY); $g.DrawString("Tc superficie cobre", $font, $black, $legendX+485, $legendY-18)
$g.DrawLine($greenPen, $legendX+880, $legendY, $legendX+960, $legendY); $g.DrawString("Ts superficie Rubatex", $font, $black, $legendX+975, $legendY-18)

$ztop=850; $zleft=145; $zW=1880; $zH=350
$g.FillRectangle($stableBrush, $zleft, $ztop, $zW, $zH)
$g.DrawRectangle($axisPen, $zleft, $ztop, $zW, $zH)
$g.DrawString("Detalle de la ventana promediada: estabilidad relativa en Ts y Tinf, con deriva residual en Tc.", $fontBold, $black, $zleft, $ztop-45)
$stable = @($data | Where-Object stable -eq 1)
$zMinM = [double]$stable[0].minute; $zMaxM = [double]$stable[-1].minute
$zMinV=30; $zMaxV=55
for($yv=30; $yv -le 55; $yv+=5){
  $yy=Y $yv $ztop $zH $zMinV $zMaxV
  $g.DrawLine($gridPen, $zleft, $yy, $zleft+$zW, $yy)
  $g.DrawString("$yv", $fontSmall, $black, $zleft-55, $yy-13)
}
function ZX($m) { return [float]($zleft + (([double]$m - $zMinM) / ($zMaxM - $zMinM)) * $zW) }
function Draw-ZSeries($key, $pen) {
  $pts = New-Object System.Collections.Generic.List[System.Drawing.PointF]
  foreach($r in $stable){
    $pts.Add([System.Drawing.PointF]::new((ZX $r.minute), (Y ([double]$r.$key) $ztop $zH $zMinV $zMaxV)))
  }
  $g.DrawLines($pen, $pts.ToArray())
}
Draw-ZSeries "amb" $bluePen
Draw-ZSeries "tc" $orangePen
Draw-ZSeries "ts" $greenPen
$g.DrawString("Fuente: elaboracion propia con datos experimentales del 29/05/2026.", $fontSmall, [System.Drawing.Brushes]::DimGray, 145, 1285)
$bmp.Save($OutPath, [System.Drawing.Imaging.ImageFormat]::Png)
$g.Dispose(); $bmp.Dispose()
''',
            encoding="utf-8-sig",
        )
        subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(ps_path),
                "-CsvPath",
                str(csv_path),
                "-OutPath",
                str(CHART_PNG),
            ],
            check=True,
        )


def text_run(text: str, bold: bool = False, italic: bool = False) -> ET.Element:
    run = ET.Element(qn("r"))
    if bold or italic:
        rpr = ET.SubElement(run, qn("rPr"))
        if bold:
            ET.SubElement(rpr, qn("b"))
        if italic:
            ET.SubElement(rpr, qn("i"))
    t = ET.SubElement(run, qn("t"))
    t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = text
    return run


def para(text: str = "", bold: bool = False, italic: bool = False, center: bool = False) -> ET.Element:
    p = ET.Element(qn("p"))
    ppr = ET.SubElement(p, qn("pPr"))
    if center:
        jc = ET.SubElement(ppr, qn("jc"))
        jc.set(qn("val"), "center")
    p.append(text_run(text, bold=bold, italic=italic))
    return p


def table(rows: list[list[str]]) -> ET.Element:
    tbl = ET.Element(qn("tbl"))
    tblpr = ET.SubElement(tbl, qn("tblPr"))
    borders = ET.SubElement(tblpr, qn("tblBorders"))
    for side in ["top", "left", "bottom", "right", "insideH", "insideV"]:
        b = ET.SubElement(borders, qn(side))
        b.set(qn("val"), "single")
        b.set(qn("sz"), "6")
        b.set(qn("space"), "0")
        b.set(qn("color"), "B7B7B7")
    for ridx, row in enumerate(rows):
        tr = ET.SubElement(tbl, qn("tr"))
        for cell in row:
            tc = ET.SubElement(tr, qn("tc"))
            tcpr = ET.SubElement(tc, qn("tcPr"))
            tcw = ET.SubElement(tcpr, qn("tcW"))
            tcw.set(qn("w"), "2400")
            tcw.set(qn("type"), "dxa")
            tc.append(para(cell, bold=(ridx == 0)))
    return tbl


def math_para(text: str) -> ET.Element:
    p = ET.Element(qn("p"))
    ppr = ET.SubElement(p, qn("pPr"))
    jc = ET.SubElement(ppr, qn("jc"))
    jc.set(qn("val"), "center")
    omath_para = ET.SubElement(p, mn("oMathPara"))
    omath = ET.SubElement(omath_para, mn("oMath"))
    run = ET.SubElement(omath, mn("r"))
    t = ET.SubElement(run, mn("t"))
    t.text = text
    return p


def replace_caption_and_append(document_xml: bytes, stats: dict[str, dict[str, float | str]]) -> bytes:
    root = ET.fromstring(document_xml)
    body = root.find(qn("body"))
    assert body is not None

    for p in body.findall(qn("p")):
        text = "".join((t.text or "") for t in p.findall(f".//{qn('t')}"))
        if text.startswith("Figura 3. Evolución temporal"):
            for child in list(p):
                p.remove(child)
            p.append(
                text_run(
                    "Figura 3. Evolución temporal de temperaturas y validación de la ventana cuasi-estacionaria. "
                    "El sombreado identifica la ventana empleada para promediar; el panel inferior amplía esa zona y permite justificar sus limitaciones.",
                    italic=True,
                )
            )
        if text.startswith("El cálculo final usa la longitud física total"):
            for child in list(p):
                p.remove(child)
            p.append(
                text_run(
                    "El cálculo final usa la longitud física total del tramo analizado, L = 0.86 m, porque esa es la longitud activa del balance de energía del segmento medido. "
                    "La ventana de datos no corresponde a un estado estacionario perfecto: se adopta como régimen cuasi-estacionario porque la superficie externa del Rubatex y el ambiente presentan variaciones pequeñas frente a la diferencia térmica de conducción, mientras que la deriva remanente de T_c se conserva como una limitación experimental explícita.",
                )
            )
        if text.startswith("Datos experimentales del proyecto Banco térmico Rubatex."):
            ref_insert_after = list(body).index(p) + 1
        if text.startswith("También es valioso mostrar que el banco"):
            discussion_insert_before = list(body).index(p)

    for t in root.findall(f".//{qn('t')}"):
        if t.text and t.text.strip() == "Catálogos técnicos":
            t.text = "Armacell, 2026"
        elif t.text and t.text.strip() == "Catálogos":
            t.text = "Armacell, 2026"
        elif t.text and t.text.strip() == "técnicos":
            t.text = ""
        elif t.text and t.text.strip() == "Rango típico de aislamiento":
            t.text = "Rango típico en espuma elastomérica flexible"

    sect_pr = body.find(qn("sectPr"))
    insert_at = list(body).index(sect_pr) if sect_pr is not None else len(list(body))

    if "ref_insert_after" in locals():
        body.insert(
            ref_insert_after,
            para(
                "Armacell. (2026). AP/ArmaFlex W technical data sheet. https://www.armacell.com/en-US/ap-armaflex-w-technical-datasheet"
            ),
        )
        if ref_insert_after <= insert_at:
            insert_at += 1

    if "discussion_insert_before" in locals():
        body.insert(
            discussion_insert_before,
            para(
                "La revisión gráfica permite precisar mejor el alcance del régimen empleado. La curva completa muestra una etapa de calentamiento y una ventana final con comportamiento más regular, pero no un estado estacionario perfecto. Por esa razón se usa la expresión régimen cuasi-estacionario: los promedios representan el tramo más estable disponible, mientras que las pendientes residuales se reportan como incertidumbre experimental y como recomendación de mejora para pruebas más largas.",
            ),
        )
        if discussion_insert_before <= insert_at:
            insert_at += 1

    w = stats["window"]
    stability_table = [
        ["Magnitud", "Media (°C)", "Desv. estándar (°C)", "Pendiente (°C/min)", "Interpretación"],
        ["T_c", f"{stats['tc']['mean']:.2f}", f"{stats['tc']['sd']:.2f}", f"{stats['tc']['slope']:.3f}", "Deriva residual; se reporta como limitación"],
        ["T_s", f"{stats['ts']['mean']:.2f}", f"{stats['ts']['sd']:.2f}", f"{stats['ts']['slope']:.3f}", "Variación moderada en la superficie externa"],
        ["T_∞", f"{stats['amb']['mean']:.2f}", f"{stats['amb']['sd']:.2f}", f"{stats['amb']['slope']:.3f}", "Ambiente prácticamente estable"],
    ]
    appendix = [
        para("Apéndice C. Validación de régimen cuasi-estacionario y desarrollo numérico completo", bold=True),
        para(
            f"La ventana de cálculo corresponde a los últimos {int(w['n'])} registros, desde {w['start']} hasta {w['end']}, con una duración aproximada de {w['duration']:.2f} min. "
            "La selección no se presenta como estado estacionario ideal, sino como una aproximación cuasi-estacionaria justificada por promedios locales y por la baja variabilidad relativa de la superficie externa del aislante.",
        ),
        table(stability_table),
        para(
            "Criterio aplicado: se usan los promedios de T_c, T_s y T_∞ de la ventana final porque el balance requiere valores representativos de la zona térmica de análisis. "
            "La pendiente positiva de T_c indica que el sistema seguía ajustándose; por ello el informe conserva esta observación como fuente de incertidumbre y recomienda extender la prueba en ensayos posteriores.",
        ),
        para("Desarrollo de ecuaciones usado para el resultado principal", bold=True),
        para("1. Temperatura de película del aire:", bold=True),
        math_para("T_f = (T_s + T_∞)/2 = (44.36 + 31.54)/2 = 37.95 °C = 311.10 K"),
        para("2. Coeficiente de expansión volumétrica para gas ideal:", bold=True),
        math_para("β = 1/T_f = 1/311.10 = 3.214×10⁻³ K⁻¹"),
        para("3. Longitud característica para convección natural en cilindro horizontal:", bold=True),
        math_para("D = 2r₂ = 2(6.75×10⁻³) = 0.0135 m"),
        para("4. Número de Rayleigh:", bold=True),
        math_para("Ra_D = gβ(T_s − T_∞)D³/(να) = 2.423×10³"),
        para("5. Número de Nusselt con la correlación de Churchill-Chu para cilindro horizontal:", bold=True),
        math_para("Nu_D = 0.60 + [0.387 Ra_D^(1/6)]/[1 + (0.559/Pr)^(9/16)]^(8/27) = 3.157"),
        para("6. Coeficiente convectivo externo:", bold=True),
        math_para("h = Nu_D k_aire/D = (3.157)(0.0271)/0.0135 = 6.34 W/m²·K"),
        para("7. Área lateral externa del Rubatex:", bold=True),
        math_para("A = 2πr₂L = 2π(0.00675)(0.86) = 0.03648 m²"),
        para("8. Calor por convección:", bold=True),
        math_para("Q_conv = hA(T_s − T_∞) = (6.34)(0.03648)(44.36 − 31.54) = 2.965 W"),
        para("9. Calor por radiación:", bold=True),
        math_para("Q_rad = εσA(T_s⁴ − T_∞⁴) = 3.035 W, usando temperaturas absolutas en K"),
        para("10. Balance externo y despeje de la conductividad:", bold=True),
        math_para("Q_total = Q_conv + Q_rad = 2.965 + 3.035 = 6.001 W"),
        math_para("k = Q_total ln(r₂/r₁)/[2πL(T_c − T_s)] = 0.0485 W/m·K"),
        para(
            "Cada fórmula anterior se basa en el balance de energía de pared cilíndrica en régimen estacionario local y en propiedades del aire evaluadas a temperatura de película, según el procedimiento de transferencia de calor reportado por Incropera, DeWitt, Bergman y Lavine (2011).",
        ),
    ]
    for offset, item in enumerate(appendix):
        body.insert(insert_at + offset, item)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def build_docx(stats: dict[str, dict[str, float | str]]) -> None:
    if OUTPUT_DOCX.exists():
        OUTPUT_DOCX.unlink()
    shutil.copyfile(SOURCE_DOCX, OUTPUT_DOCX)
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        with ZipFile(SOURCE_DOCX, "r") as zin:
            zin.extractall(tmpdir)
        shutil.copyfile(CHART_PNG, tmpdir / "word" / "media" / "image3.png")
        doc_xml_path = tmpdir / "word" / "document.xml"
        doc_xml_path.write_bytes(replace_caption_and_append(doc_xml_path.read_bytes(), stats))
        with ZipFile(OUTPUT_DOCX, "w", ZIP_DEFLATED) as zout:
            for path in tmpdir.rglob("*"):
                if path.is_file():
                    zout.write(path, path.relative_to(tmpdir).as_posix())


def main() -> None:
    rows = load_temperature_rows()
    stats = stability_stats(rows)
    write_chart_with_powershell(rows, stats)
    build_docx(stats)
    print(f"Documento generado: {OUTPUT_DOCX}")
    print(f"Figura generada: {CHART_PNG}")
    print("Ventana:", stats["window"])
    for key in ["tc", "ts", "amb"]:
        s = stats[key]
        print(f"{key}: media={s['mean']:.2f}, sd={s['sd']:.2f}, pendiente={s['slope']:.3f} C/min")


if __name__ == "__main__":
    main()
